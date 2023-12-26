import ssl
import openpyxl
from openpyxl import workbook, load_workbook
from cryptography import x509
from cryptography.hazmat.backends import default_backend
import socket
from datetime import datetime, timezone
import threading


def get_certificate_expiration(domain):
    try:
        context = ssl.create_default_context()
        with socket.create_connection((domain, 443), timeout=3) as sock:
            with context.wrap_socket(sock, server_hostname=domain) as ssock:
                cert = ssock.getpeercert(binary_form=True)
                cert_obj = x509.load_der_x509_certificate(cert, default_backend())
                expiration_date = cert_obj.not_valid_after.replace(tzinfo=timezone.utc)

            # Add debug output
            # print(f"Domain: {domain}, Expiration Date: {expiration_date}")

            return expiration_date
    except (ssl.CertificateError, socket.gaierror, socket.timeout, TimeoutError, ConnectionRefusedError,
            ConnectionResetError, OSError) as e:
        return e


def process_domain(domain, row, worksheet, domain_column, current_time):
    expiration_date = get_certificate_expiration(domain)

    if isinstance(expiration_date, datetime):
        # 转换为不带时区信息的日期对象
        expiration_date = expiration_date.replace(tzinfo=None)

        days_until_expiration = (expiration_date - current_time).days

        expiration_date_cell = worksheet.cell(row=row, column=domain_column + 1)
        expiration_date_cell.value = expiration_date

        if days_until_expiration < 30:
            red_fill = openpyxl.styles.PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            expiration_date_cell.fill = red_fill
    else:
        error_message = str(expiration_date)
        error_cell = worksheet.cell(row=row, column=domain_column + 1)
        error_cell.value = error_message
        error_fill = openpyxl.styles.PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        error_cell.fill = error_fill


def main():
    # 打开Excel文件
    workbook = openpyxl.load_workbook('domains.xlsx')

    # 选择要操作的工作表
    sheet = 'Sheet1'
    worksheet = workbook[sheet]

    # 获取域名列表，假设域名列表从第1行开始（行索引从1开始）
    domain_column = 1  # 列A
    start_row = 1

    # 计算当前时间
    current_time = datetime.now(timezone.utc).replace(tzinfo=None)

    threads = []
    # 循环遍历域名列表
    for row in range(start_row, worksheet.max_row + 1):
        domain_cell = worksheet.cell(row=row, column=domain_column)
        domain = domain_cell.value
        print(domain)

        thread = threading.Thread(target=process_domain, args=(domain, row, worksheet, domain_column, current_time))
        thread.start()
        threads.append(thread)

    # 等待所有线程完成
    for thread in threads:
        thread.join()

    # 保存Excel文件
    results_file = f'{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}_{sheet}.xlsx'
    workbook.save(results_file)

    # 打开已保存的工作簿并设置活动工作表
    saved_workbook = load_workbook(results_file)
    sheet_to_open = saved_workbook[sheet]
    saved_workbook.active = sheet_to_open
    saved_workbook.save(results_file)


if __name__ == '__main__':
    main()
